import hashlib
import io
import posixpath
import re
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import PurePosixPath
from uuid import uuid4
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile, ZipInfo

import pandas as pd
import streamlit as st
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.datetime import (
    CALENDAR_MAC_1904,
    CALENDAR_WINDOWS_1900,
    from_excel,
    to_excel,
)


# Version 1.1.0 adds chronological sorting by Visit Date and Visit Time to
# every worksheet changed during report generation.
APP_VERSION = "1.1.0"


ALCOHOL = "Alcohol"
HOME_DELIVERY = "Supermarket Home Delivery"
E_CIG = "E-Cig"
RAPID_DELIVERY = "Rapid Delivery"

SUPPORTED_TYPES = (ALCOHOL, HOME_DELIVERY, E_CIG, RAPID_DELIVERY)
TYPE_ALIASES = {
    "alcohol": ALCOHOL,
    "supermarket home delivery": HOME_DELIVERY,
    "e-cig": E_CIG,
    "e-cigarette": E_CIG,
    "rapid delivery": RAPID_DELIVERY,
}

BASE_EXPORT_COLUMNS = {
    "item_to_order",
    "site_code",
    "site_name",
    "date_of_visit_local",
    "time_of_visit_local",
    "primary_result",
}

STAFF_ASKED_ID = "Did the staff member who served you ask for ID?"
TILL_TYPE = "At which type of till was the purchase made?"
CHECKOUT_NUMBER = "What was the number of the checkout till you used?"
STAFF_NAME = "What was the name of the staff member who served you?"
STORE_RECEIPT_CODE = (
    "From the bottom of the receipt, please enter the fifteen digit code:"
)

DRIVER_ASKED_ID = "Did the driver ask for ID?"
DELIVERY_STAMPED = (
    "Has your delivery docket/receipt been stamped with the “This Delivery "
    "contains Alcohol - Please apply our Think 25 Process” message?"
)
DELIVERY_TIME = "What time did your delivery arrive?"
DRIVER_UNIFORM = "Was the driver wearing a uniform?"
DRIVER_PRESENTABLE = "Was the driver neat and presentable?"
DRIVER_NAME_BADGE = "Was the driver wearing a name badge?"
DELIVERY_THANKED = "Were you thanked at any stage?"
DELIVERY_FRIENDLY = (
    "Overall, was the service that you received helpful and friendly?"
)
HOME_ORDER_REFERENCE = "Please enter the order number:"
RAPID_ORDER_REFERENCE = (
    "Please enter the order number from your online receipt "
    "(including any special characters):"
)

STORE_HEADERS = [
    "Store Number",
    "Name",
    "Visit Date",
    "Visit Time",
    "Result",
    "Asked for ID",
    "Checkout Number",
    "Name Tag",
    "Receipt Code",
    "Till (If Fail)",
]

VAPE_HEADERS = STORE_HEADERS[:9]

DOT_COM_HEADERS = [
    "Store Number",
    "Name",
    "Visit Date",
    "Visit Time",
    "Result",
    "Asked for ID",
    "Delivery receipt stamped?",
    "Time of Delivery",
    "Driver in Uniform",
    "Driver Neat and Presentable",
    "Driver Wearing a Name Tag",
    "Were You Thanked",
    "Friendly & Helpful Service",
    "Order Reference",
]

WHOOSH_HEADERS = [
    "Store Number",
    "Name",
    "Visit Date",
    "Visit Time",
    "Result",
    "Asked for ID",
    "Receipt Code",
    "First or Second?",
]

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
EXT_PROPS_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
)
VT_NS = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

PARSER = etree.XMLParser(remove_blank_text=False, resolve_entities=False)


@dataclass
class GenerationResult:
    main_report: bytes
    whoosh_report: bytes
    main_filename: str
    whoosh_filename: str
    reporting_week: int
    stats: dict
    ignored_rows: int


def clean_text(value):
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    return str(value).strip()


def normalise_header(value):
    return re.sub(r"\s+", " ", clean_text(value)).casefold()


def normalise_sheet_name(value):
    return re.sub(r"\s+", " ", clean_text(value)).casefold()


def parse_store_number(value):
    match = re.match(r"^\s*(\d+)", clean_text(value))
    if not match:
        return None
    return int(match.group(1))


def parse_date_value(value, epoch=CALENDAR_WINDOWS_1900):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch=epoch)
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        return None

    text = clean_text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    converted = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return None if pd.isna(converted) else converted.date()


def parse_time_value(value, epoch=CALENDAR_WINDOWS_1900):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return time(value.hour, value.minute)
    if isinstance(value, time):
        return time(value.hour, value.minute)
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch=epoch)
        if isinstance(converted, datetime):
            converted = converted.time()
        if isinstance(converted, time):
            return time(converted.hour, converted.minute)
        return None

    text = clean_text(value)
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
        try:
            converted = datetime.strptime(text, fmt)
            return time(converted.hour, converted.minute)
        except ValueError:
            pass
    return None


def visit_key(store_number, visit_date, visit_time):
    if store_number is None or visit_date is None or visit_time is None:
        return None
    return (
        int(store_number),
        visit_date,
        visit_time.hour,
        visit_time.minute,
    )


def numeric_if_digits(value, preserve_leading_zero=False):
    text = clean_text(value)
    if not text:
        return None
    if (
        text.isdigit()
        and len(text) <= 15
        and not (preserve_leading_zero and len(text) > 1 and text[0] == "0")
    ):
        return int(text)
    return text


def row_identifier(row, fallback_index):
    internal_id = clean_text(row.get("internal_id"))
    return internal_id or f"CSV row {fallback_index + 2}"


def read_export(export_bytes):
    try:
        df = pd.read_csv(
            io.BytesIO(export_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        df = pd.read_csv(
            io.BytesIO(export_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="cp1252",
        )

    df.columns = [clean_text(column) for column in df.columns]
    missing = sorted(BASE_EXPORT_COLUMNS - set(df.columns))
    if missing:
        raise KeyError("The export is missing required column(s): " + ", ".join(missing))

    df["_audit_type"] = (
        df["item_to_order"].map(lambda value: TYPE_ALIASES.get(clean_text(value).casefold()))
    )
    supported = df[df["_audit_type"].notna()].copy()
    ignored_rows = len(df) - len(supported)
    if supported.empty:
        raise ValueError(
            "The export does not contain Alcohol, Supermarket Home Delivery, "
            "E-Cig or Rapid Delivery audits."
        )

    present_types = set(supported["_audit_type"])
    required_questions = set()
    if present_types & {ALCOHOL, E_CIG}:
        required_questions.update(
            {
                STAFF_ASKED_ID,
                CHECKOUT_NUMBER,
                STAFF_NAME,
                STORE_RECEIPT_CODE,
            }
        )
    if ALCOHOL in present_types:
        required_questions.add(TILL_TYPE)
    if HOME_DELIVERY in present_types:
        required_questions.update(
            {
                DRIVER_ASKED_ID,
                DELIVERY_STAMPED,
                DELIVERY_TIME,
                DRIVER_UNIFORM,
                DRIVER_PRESENTABLE,
                DRIVER_NAME_BADGE,
                DELIVERY_THANKED,
                DELIVERY_FRIENDLY,
                HOME_ORDER_REFERENCE,
            }
        )
    if RAPID_DELIVERY in present_types:
        required_questions.update({DRIVER_ASKED_ID, RAPID_ORDER_REFERENCE})

    missing_questions = sorted(required_questions - set(df.columns))
    if missing_questions:
        raise KeyError(
            "The export is missing required question column(s): "
            + "; ".join(missing_questions)
        )

    invalid = []
    stores = []
    dates = []
    times = []
    for index, row in supported.iterrows():
        store_number = parse_store_number(row["site_code"])
        visit_date = parse_date_value(row["date_of_visit_local"])
        visit_time = parse_time_value(row["time_of_visit_local"])
        if store_number is None or visit_date is None or visit_time is None:
            invalid.append(row_identifier(row, index))
        stores.append(store_number)
        dates.append(visit_date)
        times.append(visit_time)

    if invalid:
        preview = ", ".join(invalid[:12])
        extra = "" if len(invalid) <= 12 else f" and {len(invalid) - 12} more"
        raise ValueError(
            "Store Number, local visit date or local visit time could not be read for: "
            + preview
            + extra
        )

    supported["_store_number"] = stores
    supported["_visit_date"] = dates
    supported["_visit_time"] = times
    return supported, ignored_rows


def read_calendar(calendar_bytes):
    workbook = load_workbook(
        io.BytesIO(calendar_bytes), read_only=True, data_only=True
    )
    candidates = []
    try:
        for worksheet in workbook.worksheets:
            max_column = min(worksheet.max_column, 60)
            for header_row in range(1, min(worksheet.max_row, 12) + 1):
                values = next(
                    worksheet.iter_rows(
                        min_row=header_row,
                        max_row=header_row,
                        max_col=max_column,
                        values_only=True,
                    )
                )
                positions = {}
                for column_number, value in enumerate(values, start=1):
                    header = re.sub(r"[^a-z0-9]", "", normalise_header(value))
                    if header in {"date"}:
                        positions.setdefault("date", column_number)
                    elif header in {"pd", "period"}:
                        positions.setdefault("period", column_number)
                    elif header in {"wk", "week"}:
                        positions.setdefault("week", column_number)

                if set(positions) != {"date", "period", "week"}:
                    continue

                mapping = {}
                for row in worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=worksheet.max_row,
                    max_col=max(positions.values()),
                    values_only=True,
                ):
                    visit_date = parse_date_value(
                        row[positions["date"] - 1], workbook.epoch
                    )
                    if visit_date is None:
                        continue
                    try:
                        period = int(float(row[positions["period"] - 1]))
                        week = int(float(row[positions["week"] - 1]))
                    except (TypeError, ValueError):
                        continue
                    if not 1 <= period <= 12 or week < 1:
                        continue
                    if visit_date in mapping and mapping[visit_date] != (period, week):
                        raise ValueError(
                            f"The calendar assigns {visit_date:%d/%m/%Y} to more than one period or week."
                        )
                    mapping[visit_date] = (period, week)

                if mapping:
                    candidates.append(mapping)
    finally:
        workbook.close()

    if not candidates:
        raise ValueError(
            "Could not find a calendar table containing Date, Pd and Wk columns."
        )
    return max(candidates, key=len)


class OOXMLWorkbook:
    """Targeted XLSX editor that preserves unsupported Excel parts byte-for-byte."""

    def __init__(self, workbook_bytes, label):
        self.label = label
        try:
            with ZipFile(io.BytesIO(workbook_bytes), "r") as archive:
                self.original_infos = archive.infolist()
                self.parts = {
                    info.filename: archive.read(info.filename)
                    for info in self.original_infos
                }
        except BadZipFile as exc:
            raise ValueError(f"{label} is not a valid .xlsx file.") from exc

        required = {
            "[Content_Types].xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
        }
        if not required.issubset(self.parts):
            raise ValueError(f"{label} is missing required Excel workbook parts.")

        self.xml_roots = {}
        self.new_part_infos = []
        self.shared_strings = self._read_shared_strings()
        self._refresh_sheet_map()

        workbook_pr = self.workbook_root.find(f"{{{MAIN_NS}}}workbookPr")
        uses_1904 = workbook_pr is not None and workbook_pr.get("date1904") in {
            "1",
            "true",
            "True",
        }
        self.epoch = CALENDAR_MAC_1904 if uses_1904 else CALENDAR_WINDOWS_1900

    def _parse_part(self, part_name):
        if part_name not in self.xml_roots:
            self.xml_roots[part_name] = etree.fromstring(
                self.parts[part_name], parser=PARSER
            )
        return self.xml_roots[part_name]

    @property
    def workbook_root(self):
        return self._parse_part("xl/workbook.xml")

    @property
    def workbook_rels_root(self):
        return self._parse_part("xl/_rels/workbook.xml.rels")

    def _read_shared_strings(self):
        part_name = "xl/sharedStrings.xml"
        if part_name not in self.parts:
            return []
        root = etree.fromstring(self.parts[part_name], parser=PARSER)
        values = []
        for item in root.findall(f"{{{MAIN_NS}}}si"):
            values.append("".join(item.itertext()))
        return values

    def _refresh_sheet_map(self):
        relationships = {
            relationship.get("Id"): relationship.get("Target")
            for relationship in self.workbook_rels_root.findall(
                f"{{{PKG_REL_NS}}}Relationship"
            )
        }
        self.sheet_map = {}
        sheets = self.workbook_root.find(f"{{{MAIN_NS}}}sheets")
        for sheet in sheets.findall(f"{{{MAIN_NS}}}sheet"):
            relationship_id = sheet.get(f"{{{DOC_REL_NS}}}id")
            target = relationships.get(relationship_id)
            if not target:
                continue
            if target.startswith("/"):
                part_name = target.lstrip("/")
            else:
                part_name = posixpath.normpath(posixpath.join("xl", target))
            self.sheet_map[sheet.get("name")] = part_name

    @property
    def sheet_names(self):
        return list(self.sheet_map)

    def find_sheet(self, logical_name):
        wanted = normalise_sheet_name(logical_name)
        matches = [
            name for name in self.sheet_names if normalise_sheet_name(name) == wanted
        ]
        if len(matches) > 1:
            raise ValueError(
                f"{self.label} contains more than one sheet matching {logical_name!r}."
            )
        return matches[0] if matches else None

    def sheet_root(self, sheet_name):
        return self._parse_part(self.sheet_map[sheet_name])

    @staticmethod
    def _sheet_data(root):
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            raise ValueError("A report worksheet does not contain sheet data.")
        return sheet_data

    @staticmethod
    def _row_number(row):
        return int(row.get("r"))

    @staticmethod
    def _column_number(cell):
        match = re.match(r"([A-Z]+)", cell.get("r", ""))
        if not match:
            return None
        number = 0
        for character in match.group(1):
            number = number * 26 + ord(character) - 64
        return number

    def rows(self, sheet_name):
        return self._sheet_data(self.sheet_root(sheet_name)).findall(
            f"{{{MAIN_NS}}}row"
        )

    def row_by_number(self, sheet_name, row_number):
        for row in self.rows(sheet_name):
            if self._row_number(row) == row_number:
                return row
        return None

    def _ensure_row(self, sheet_name, row_number, source_row=None):
        root = self.sheet_root(sheet_name)
        sheet_data = self._sheet_data(root)
        existing = self.row_by_number(sheet_name, row_number)
        if existing is not None:
            return existing

        attributes = {}
        if source_row is not None:
            attributes = {
                key: value
                for key, value in source_row.attrib.items()
                if key not in {"r", "spans"}
            }
        attributes["r"] = str(row_number)
        row = etree.Element(f"{{{MAIN_NS}}}row", attributes)

        inserted = False
        for position, other in enumerate(sheet_data):
            if self._row_number(other) > row_number:
                sheet_data.insert(position, row)
                inserted = True
                break
        if not inserted:
            sheet_data.append(row)
        return row

    def cell_in_row(self, row, column_number):
        for cell in row.findall(f"{{{MAIN_NS}}}c"):
            if self._column_number(cell) == column_number:
                return cell
        return None

    def _ensure_cell(self, row, column_number, style_source=None):
        existing = self.cell_in_row(row, column_number)
        if existing is not None:
            if style_source is not None and style_source.get("s") is not None:
                existing.set("s", style_source.get("s"))
            return existing

        reference = f"{get_column_letter(column_number)}{self._row_number(row)}"
        cell = etree.Element(f"{{{MAIN_NS}}}c", {"r": reference})
        if style_source is not None and style_source.get("s") is not None:
            cell.set("s", style_source.get("s"))

        inserted = False
        for position, other in enumerate(row.findall(f"{{{MAIN_NS}}}c")):
            other_column = self._column_number(other)
            if other_column is not None and other_column > column_number:
                row.insert(position, cell)
                inserted = True
                break
        if not inserted:
            row.append(cell)
        return cell

    def cell_value(self, cell):
        if cell is None:
            return None
        cell_type = cell.get("t")
        if cell_type == "inlineStr":
            inline = cell.find(f"{{{MAIN_NS}}}is")
            return None if inline is None else "".join(inline.itertext())

        value = cell.find(f"{{{MAIN_NS}}}v")
        if value is None or value.text is None:
            return None
        text = value.text
        if cell_type == "s":
            try:
                return self.shared_strings[int(text)]
            except (ValueError, IndexError):
                return None
        if cell_type == "b":
            return text == "1"
        if cell_type in {"str", "e"}:
            return text
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return text

    def value_at(self, sheet_name, row_number, column_number):
        row = self.row_by_number(sheet_name, row_number)
        return None if row is None else self.cell_value(
            self.cell_in_row(row, column_number)
        )

    def _write_value(self, cell, value):
        for child in list(cell):
            cell.remove(child)
        cell.attrib.pop("t", None)

        if value is None or value == "":
            return
        if isinstance(value, datetime):
            numeric = to_excel(value, epoch=self.epoch)
            value_element = etree.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = format(numeric, ".15g")
            return
        if isinstance(value, date):
            numeric = to_excel(datetime.combine(value, time()), epoch=self.epoch)
            value_element = etree.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = format(numeric, ".15g")
            return
        if isinstance(value, time):
            seconds = value.hour * 3600 + value.minute * 60 + value.second
            numeric = seconds / 86400
            value_element = etree.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = format(numeric, ".15g")
            return
        if isinstance(value, bool):
            cell.set("t", "b")
            value_element = etree.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = "1" if value else "0"
            return
        if isinstance(value, (int, float)):
            value_element = etree.SubElement(cell, f"{{{MAIN_NS}}}v")
            value_element.text = str(value)
            return

        text = str(value)
        cell.set("t", "inlineStr")
        inline = etree.SubElement(cell, f"{{{MAIN_NS}}}is")
        text_element = etree.SubElement(inline, f"{{{MAIN_NS}}}t")
        if text != text.strip() or "  " in text:
            text_element.set(XML_SPACE, "preserve")
        text_element.text = text

    def _update_dimension(self, sheet_name, row_number, column_number):
        root = self.sheet_root(sheet_name)
        dimension = root.find(f"{{{MAIN_NS}}}dimension")
        if dimension is None:
            dimension = etree.Element(
                f"{{{MAIN_NS}}}dimension", {"ref": "A1"}
            )
            root.insert(0, dimension)
        try:
            min_col, min_row, max_col, max_row = range_boundaries(
                dimension.get("ref", "A1")
            )
        except ValueError:
            min_col = min_row = max_col = max_row = 1
        max_col = max(max_col, column_number)
        max_row = max(max_row, row_number)
        dimension.set(
            "ref",
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}",
        )

    def header_values(self, sheet_name, column_count):
        return [
            self.value_at(sheet_name, 1, column_number)
            for column_number in range(1, column_count + 1)
        ]

    def header_matches(self, sheet_name, headers):
        actual = self.header_values(sheet_name, len(headers))
        return [normalise_header(value) for value in actual] == [
            normalise_header(value) for value in headers
        ]

    def ensure_headers(self, sheet_name, headers, template_sheet):
        actual = self.header_values(sheet_name, len(headers))
        populated = [value for value in actual if clean_text(value)]
        if populated:
            if [normalise_header(value) for value in actual] != [
                normalise_header(value) for value in headers
            ]:
                raise ValueError(
                    f"The headers in {self.label} sheet {sheet_name!r} do not match the expected layout."
                )
            return

        template_row = self.row_by_number(template_sheet, 1)
        target_row = self._ensure_row(sheet_name, 1, template_row)
        for column_number, header in enumerate(headers, start=1):
            source_cell = (
                None
                if template_row is None
                else self.cell_in_row(template_row, column_number)
            )
            target_cell = self._ensure_cell(target_row, column_number, source_cell)
            self._write_value(target_cell, header)
            self._update_dimension(sheet_name, 1, column_number)

    def last_data_row(self, sheet_name):
        last = 1
        for row in self.rows(sheet_name):
            row_number = self._row_number(row)
            if row_number < 2:
                continue
            value = self.cell_value(self.cell_in_row(row, 1))
            if value not in (None, ""):
                last = max(last, row_number)
        return last

    def append_values(self, sheet_name, values, template_sheet):
        last_row = self.last_data_row(sheet_name)
        target_row_number = last_row + 1
        if last_row >= 2:
            style_row = self.row_by_number(sheet_name, last_row)
        else:
            style_row = self.row_by_number(template_sheet, 2)

        target_row = self._ensure_row(sheet_name, target_row_number, style_row)
        template_style_row = self.row_by_number(template_sheet, 2)
        for column_number, value in enumerate(values, start=1):
            style_cell = (
                None
                if style_row is None
                else self.cell_in_row(style_row, column_number)
            )
            if style_cell is None and template_style_row is not None:
                style_cell = self.cell_in_row(template_style_row, column_number)
            target_cell = self._ensure_cell(target_row, column_number, style_cell)
            self._write_value(target_cell, value)
            self._update_dimension(sheet_name, target_row_number, column_number)

        if target_row.get("spans") is not None:
            target_row.set("spans", f"1:{len(values)}")
        return target_row_number

    def sort_data_rows(self, sheet_name, column_count):
        """Sort populated data rows by Visit Date and Visit Time, oldest first."""
        records = []
        populated_row_numbers = []
        for row in self.rows(sheet_name):
            row_number = self._row_number(row)
            if row_number < 2:
                continue
            store_number = self.cell_value(self.cell_in_row(row, 1))
            if store_number in (None, ""):
                continue

            values = [
                self.cell_value(self.cell_in_row(row, column_number))
                for column_number in range(1, column_count + 1)
            ]
            visit_date = parse_date_value(values[2], self.epoch)
            visit_time = parse_time_value(values[3], self.epoch)
            records.append(
                (visit_date, visit_time, row_number, values)
            )
            populated_row_numbers.append(row_number)

        if not records:
            return

        records.sort(
            key=lambda record: (
                record[0] is None,
                record[0] or date.max,
                record[1] is None,
                24 if record[1] is None else record[1].hour,
                60 if record[1] is None else record[1].minute,
                record[2],
            )
        )

        style_row = self.row_by_number(sheet_name, 2)
        if style_row is None:
            style_row = self.row_by_number(sheet_name, populated_row_numbers[0])

        last_original_row = max(populated_row_numbers)
        for row_number in range(2, last_original_row + 1):
            row = self.row_by_number(sheet_name, row_number)
            if row is None:
                continue
            for column_number in range(1, column_count + 1):
                cell = self.cell_in_row(row, column_number)
                if cell is not None:
                    self._write_value(cell, None)

        for target_row_number, (_, _, _, values) in enumerate(records, start=2):
            target_row = self._ensure_row(
                sheet_name, target_row_number, style_row
            )
            for column_number, value in enumerate(values, start=1):
                target_cell = self.cell_in_row(target_row, column_number)
                if target_cell is None:
                    style_cell = (
                        None
                        if style_row is None
                        else self.cell_in_row(style_row, column_number)
                    )
                    target_cell = self._ensure_cell(
                        target_row, column_number, style_cell
                    )
                self._write_value(target_cell, value)
                self._update_dimension(
                    sheet_name, target_row_number, column_number
                )

    def existing_visit_keys(self, sheet_filter):
        keys = set()
        for sheet_name in self.sheet_names:
            if not sheet_filter(sheet_name):
                continue
            for row in self.rows(sheet_name):
                row_number = self._row_number(row)
                if row_number < 2:
                    continue
                store_number = parse_store_number(
                    self.cell_value(self.cell_in_row(row, 1))
                )
                visit_date = parse_date_value(
                    self.cell_value(self.cell_in_row(row, 3)), self.epoch
                )
                visit_time = parse_time_value(
                    self.cell_value(self.cell_in_row(row, 4)), self.epoch
                )
                key = visit_key(store_number, visit_date, visit_time)
                if key is not None:
                    keys.add(key)
        return keys

    def recalculate_first_or_second(self, sheet_name):
        visits_by_store = {}
        row_lookup = {}
        for row in self.rows(sheet_name):
            row_number = self._row_number(row)
            if row_number < 2:
                continue
            store_number = parse_store_number(
                self.cell_value(self.cell_in_row(row, 1))
            )
            visit_date = parse_date_value(
                self.cell_value(self.cell_in_row(row, 3)), self.epoch
            )
            visit_time = parse_time_value(
                self.cell_value(self.cell_in_row(row, 4)), self.epoch
            )
            if store_number is None or visit_date is None or visit_time is None:
                continue
            visits_by_store.setdefault(store_number, []).append(
                (visit_date, visit_time.hour, visit_time.minute, row_number)
            )
            row_lookup[row_number] = row

        for visits in visits_by_store.values():
            visits.sort()
            for position, (_, _, _, row_number) in enumerate(visits):
                row = row_lookup[row_number]
                cell = self._ensure_cell(row, 8)
                self._write_value(cell, "First" if position == 0 else "Second")
                self._update_dimension(sheet_name, row_number, 8)

    def add_sheet_by_cloning(self, new_name, template_name):
        if self.find_sheet(new_name):
            return self.find_sheet(new_name)

        cloned_root = deepcopy(self.sheet_root(template_name))
        for attribute_name in list(cloned_root.attrib):
            if etree.QName(attribute_name).localname == "uid":
                cloned_root.set(attribute_name, "{" + str(uuid4()).upper() + "}")

        sheet_data = self._sheet_data(cloned_root)
        for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
            if self._row_number(row) < 2:
                continue
            for cell in row.findall(f"{{{MAIN_NS}}}c"):
                for child in list(cell):
                    cell.remove(child)
                cell.attrib.pop("t", None)

        sheet_view = cloned_root.find(f"{{{MAIN_NS}}}sheetViews/{{{MAIN_NS}}}sheetView")
        if sheet_view is not None:
            sheet_view.attrib.pop("tabSelected", None)

        worksheet_numbers = []
        for part_name in self.parts:
            match = re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", part_name)
            if match:
                worksheet_numbers.append(int(match.group(1)))
        new_number = max(worksheet_numbers, default=0) + 1
        new_part = f"xl/worksheets/sheet{new_number}.xml"
        self.parts[new_part] = b""
        self.xml_roots[new_part] = cloned_root

        info = ZipInfo(new_part, date_time=(1980, 1, 1, 0, 0, 0))
        info.compress_type = ZIP_DEFLATED
        info.external_attr = 0o600 << 16
        self.new_part_infos.append(info)

        relationship_ids = []
        for relationship in self.workbook_rels_root.findall(
            f"{{{PKG_REL_NS}}}Relationship"
        ):
            match = re.fullmatch(r"rId(\d+)", relationship.get("Id", ""))
            if match:
                relationship_ids.append(int(match.group(1)))
        relationship_id = f"rId{max(relationship_ids, default=0) + 1}"
        etree.SubElement(
            self.workbook_rels_root,
            f"{{{PKG_REL_NS}}}Relationship",
            {
                "Id": relationship_id,
                "Type": (
                    "http://schemas.openxmlformats.org/officeDocument/2006/"
                    "relationships/worksheet"
                ),
                "Target": f"worksheets/sheet{new_number}.xml",
            },
        )

        sheets = self.workbook_root.find(f"{{{MAIN_NS}}}sheets")
        sheet_ids = [int(sheet.get("sheetId")) for sheet in sheets]
        etree.SubElement(
            sheets,
            f"{{{MAIN_NS}}}sheet",
            {
                "name": new_name,
                "sheetId": str(max(sheet_ids, default=0) + 1),
                f"{{{DOC_REL_NS}}}id": relationship_id,
            },
        )

        content_types = self._parse_part("[Content_Types].xml")
        etree.SubElement(
            content_types,
            f"{{{CONTENT_TYPES_NS}}}Override",
            {
                "PartName": f"/{new_part}",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.worksheet+xml"
                ),
            },
        )
        self._update_extended_properties(new_name)
        self._refresh_sheet_map()
        return new_name

    def _update_extended_properties(self, new_sheet_name):
        part_name = "docProps/app.xml"
        if part_name not in self.parts:
            return
        root = self._parse_part(part_name)
        titles_vector = root.find(
            f"{{{EXT_PROPS_NS}}}TitlesOfParts/{{{VT_NS}}}vector"
        )
        if titles_vector is not None:
            etree.SubElement(
                titles_vector, f"{{{VT_NS}}}lpstr"
            ).text = new_sheet_name
            titles_vector.set("size", str(len(titles_vector)))

        heading_vector = root.find(
            f"{{{EXT_PROPS_NS}}}HeadingPairs/{{{VT_NS}}}vector"
        )
        if heading_vector is not None:
            for integer in heading_vector.findall(f".//{{{VT_NS}}}i4"):
                integer.text = str(len(self.sheet_map) + 1)
                break

    def _set_recalculation(self):
        calculation = self.workbook_root.find(f"{{{MAIN_NS}}}calcPr")
        if calculation is None:
            calculation = etree.SubElement(
                self.workbook_root, f"{{{MAIN_NS}}}calcPr"
            )
        calculation.set("calcMode", "auto")
        calculation.set("fullCalcOnLoad", "1")
        calculation.set("forceFullCalc", "1")

    def to_bytes(self):
        self._set_recalculation()
        for part_name, root in self.xml_roots.items():
            self.parts[part_name] = etree.tostring(
                root,
                xml_declaration=True,
                encoding="UTF-8",
                standalone=True,
            )

        output = io.BytesIO()
        original_names = set()
        with ZipFile(output, "w") as archive:
            for info in self.original_infos:
                original_names.add(info.filename)
                archive.writestr(info, self.parts[info.filename])
            for info in self.new_part_infos:
                if info.filename not in original_names:
                    archive.writestr(info, self.parts[info.filename])
        return output.getvalue()


def period_from_sheet_name(sheet_name, suffix=None):
    if suffix is None:
        pattern = r"P(\d+)"
    else:
        pattern = rf"P(\d+)\s+{re.escape(suffix)}"
    match = re.fullmatch(pattern + r"\s*", clean_text(sheet_name), flags=re.I)
    return None if match is None else int(match.group(1))


def main_detail_sheet(sheet_name):
    return any(
        period_from_sheet_name(sheet_name, suffix) is not None
        for suffix in ("Store Details", "Dot Com Details", "Vape Details")
    )


def whoosh_period_sheet(sheet_name):
    return period_from_sheet_name(sheet_name) is not None


def find_main_template(workbook, suffix, headers, target_period):
    candidates = []
    for sheet_name in workbook.sheet_names:
        period = period_from_sheet_name(sheet_name, suffix)
        if period is None or not workbook.header_matches(sheet_name, headers):
            continue
        preference = 1 if period <= target_period else 0
        candidates.append((preference, period, sheet_name))
    if not candidates:
        raise ValueError(
            f"{workbook.label} does not contain a populated {suffix!r} layout to use as a template."
        )
    return max(candidates)[2]


def get_main_period_sheet(workbook, period, suffix):
    sheet_name = workbook.find_sheet(f"P{period} {suffix}")
    if sheet_name is None:
        raise ValueError(
            f"{workbook.label} does not contain the required P{period} {suffix} sheet."
        )
    return sheet_name


def get_whoosh_period_sheet(workbook, period):
    wanted = f"P{period}"
    existing = workbook.find_sheet(wanted)
    if existing is not None:
        return existing

    candidates = []
    for sheet_name in workbook.sheet_names:
        sheet_period = period_from_sheet_name(sheet_name)
        if sheet_period is None or not workbook.header_matches(
            sheet_name, WHOOSH_HEADERS
        ):
            continue
        candidates.append(
            (workbook.last_data_row(sheet_name), sheet_period, sheet_name)
        )
    if not candidates:
        raise ValueError(
            f"{workbook.label} does not contain a Whoosh period sheet to use as a template."
        )
    template = max(candidates)[2]
    return workbook.add_sheet_by_cloning(wanted, template)


def map_store_row(row, include_till):
    result = clean_text(row["primary_result"]).lower()
    values = [
        int(row["_store_number"]),
        clean_text(row["site_name"]),
        row["_visit_date"],
        row["_visit_time"],
        result,
        clean_text(row[STAFF_ASKED_ID]) or None,
        numeric_if_digits(row[CHECKOUT_NUMBER]),
        clean_text(row[STAFF_NAME]) or None,
        numeric_if_digits(row[STORE_RECEIPT_CODE], preserve_leading_zero=True),
    ]
    if include_till:
        values.append(clean_text(row[TILL_TYPE]) if result == "fail" else None)
    return values


def map_dot_com_row(row):
    return [
        int(row["_store_number"]),
        clean_text(row["site_name"]),
        row["_visit_date"],
        row["_visit_time"],
        clean_text(row["primary_result"]).lower(),
        clean_text(row[DRIVER_ASKED_ID]) or None,
        clean_text(row[DELIVERY_STAMPED]) or None,
        parse_time_value(row[DELIVERY_TIME]),
        clean_text(row[DRIVER_UNIFORM]) or None,
        clean_text(row[DRIVER_PRESENTABLE]) or None,
        clean_text(row[DRIVER_NAME_BADGE]) or None,
        clean_text(row[DELIVERY_THANKED]) or None,
        clean_text(row[DELIVERY_FRIENDLY]) or None,
        numeric_if_digits(row[HOME_ORDER_REFERENCE], preserve_leading_zero=True),
    ]


def map_whoosh_row(row):
    return [
        int(row["_store_number"]),
        clean_text(row["site_name"]),
        row["_visit_date"],
        row["_visit_time"],
        clean_text(row["primary_result"]).lower(),
        clean_text(row[DRIVER_ASKED_ID]) or None,
        numeric_if_digits(row[RAPID_ORDER_REFERENCE], preserve_leading_zero=True),
        None,
    ]


def updated_report_filename(original_name, reporting_week):
    name = PurePosixPath(original_name or "report.xlsx").name
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    stem = name[:-5]
    replacement = f"(Week {reporting_week})"
    updated, count = re.subn(
        r"\(\s*Week\s*\d+\s*\)", replacement, stem, flags=re.I
    )
    if count == 0:
        updated, count = re.subn(
            r"\bWeek\s*\d+\b", f"Week {reporting_week}", stem, flags=re.I
        )
    if count == 0:
        updated = f"{stem} {replacement}"
    return updated + ".xlsx"


def generate_reports(
    export_bytes,
    calendar_bytes,
    main_report_bytes,
    whoosh_report_bytes,
    main_original_name="Test Purchase Report.xlsx",
    whoosh_original_name="Test Purchase Report Whoosh.xlsx",
):
    export, ignored_rows = read_export(export_bytes)
    calendar = read_calendar(calendar_bytes)

    missing_dates = sorted(
        {visit_date for visit_date in export["_visit_date"] if visit_date not in calendar}
    )
    if missing_dates:
        date_list = ", ".join(value.strftime("%d/%m/%Y") for value in missing_dates)
        raise ValueError("The following visit date(s) are missing from the calendar: " + date_list)

    export["_period"] = export["_visit_date"].map(
        lambda value: calendar[value][0]
    )
    export["_week"] = export["_visit_date"].map(lambda value: calendar[value][1])
    reporting_week = int(export["_week"].max())

    main = OOXMLWorkbook(main_report_bytes, "the latest main report")
    whoosh = OOXMLWorkbook(whoosh_report_bytes, "the latest Whoosh report")

    stats = {
        audit_type: {"Added": 0, "Already present": 0}
        for audit_type in SUPPORTED_TYPES
    }
    main_seen = main.existing_visit_keys(main_detail_sheet)
    whoosh_seen = whoosh.existing_visit_keys(whoosh_period_sheet)
    touched_main_sheets = {}
    touched_whoosh_periods = set()

    for _, row in export.iterrows():
        audit_type = row["_audit_type"]
        key = visit_key(
            row["_store_number"], row["_visit_date"], row["_visit_time"]
        )
        period = int(row["_period"])

        if audit_type == RAPID_DELIVERY:
            if key in whoosh_seen:
                stats[audit_type]["Already present"] += 1
                continue
            sheet_name = get_whoosh_period_sheet(whoosh, period)
            template_sheet = sheet_name
            whoosh.ensure_headers(sheet_name, WHOOSH_HEADERS, template_sheet)
            whoosh.append_values(
                sheet_name, map_whoosh_row(row), template_sheet
            )
            whoosh_seen.add(key)
            touched_whoosh_periods.add(period)
            stats[audit_type]["Added"] += 1
            continue

        if key in main_seen:
            stats[audit_type]["Already present"] += 1
            continue

        if audit_type == ALCOHOL:
            suffix = "Store Details"
            headers = STORE_HEADERS
            values = map_store_row(row, include_till=True)
        elif audit_type == HOME_DELIVERY:
            suffix = "Dot Com Details"
            headers = DOT_COM_HEADERS
            values = map_dot_com_row(row)
        elif audit_type == E_CIG:
            suffix = "Vape Details"
            headers = VAPE_HEADERS
            values = map_store_row(row, include_till=False)
        else:
            continue

        sheet_name = get_main_period_sheet(main, period, suffix)
        template_sheet = find_main_template(main, suffix, headers, period)
        main.ensure_headers(sheet_name, headers, template_sheet)
        main.append_values(sheet_name, values, template_sheet)
        touched_main_sheets[sheet_name] = len(headers)
        main_seen.add(key)
        stats[audit_type]["Added"] += 1

    for sheet_name, column_count in touched_main_sheets.items():
        main.sort_data_rows(sheet_name, column_count)

    for period in touched_whoosh_periods:
        sheet_name = get_whoosh_period_sheet(whoosh, period)
        whoosh.sort_data_rows(sheet_name, len(WHOOSH_HEADERS))
        whoosh.recalculate_first_or_second(sheet_name)

    return GenerationResult(
        main_report=main.to_bytes(),
        whoosh_report=whoosh.to_bytes(),
        main_filename=updated_report_filename(main_original_name, reporting_week),
        whoosh_filename=updated_report_filename(
            whoosh_original_name, reporting_week
        ),
        reporting_week=reporting_week,
        stats=stats,
        ignored_rows=ignored_rows,
    )


def make_download_zip(result):
    output = io.BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr(result.main_filename, result.main_report)
        archive.writestr(result.whoosh_filename, result.whoosh_report)
    return output.getvalue()


def main():
    st.set_page_config(
        page_title="Tesco Ireland Weekly Report Generator",
        layout="centered",
    )
    st.title("Tesco Ireland Weekly Report Generator")
    st.caption(
        f"Version {APP_VERSION} — chronological sorting enabled for updated tabs"
    )
    st.write(
        "Upload the latest audits export, Tesco calendar and the most recent "
        "main and Whoosh reports. The app will add new visits to the correct "
        "period tabs and leave visits already reported unchanged."
    )

    st.markdown(
        """
The app will:

- map Alcohol audits to **Px Store Details**
- map Supermarket Home Delivery audits to **Px Dot Com Details**
- map E-Cig audits to **Px Vape Details**
- map Rapid Delivery audits to the Whoosh **Px** tabs
- de-duplicate visits using Store Number, local visit date and local visit time
- sort each updated tab chronologically by visit date and visit time
- set Whoosh visits to **First** or **Second** within each store and period
"""
    )

    export_file = st.file_uploader(
        "Upload audits_basic_data_export.csv", type=["csv"]
    )
    calendar_file = st.file_uploader("Upload Tesco Calendar", type=["xlsx"])
    main_report_file = st.file_uploader(
        "Upload the most recent Test Purchase Report", type=["xlsx"]
    )
    whoosh_report_file = st.file_uploader(
        "Upload the most recent Whoosh Report", type=["xlsx"]
    )

    uploads = [
        export_file,
        calendar_file,
        main_report_file,
        whoosh_report_file,
    ]
    if not all(uploads):
        return

    upload_bytes = [uploaded.getvalue() for uploaded in uploads]
    signature = tuple(
        (
            uploaded.name,
            len(content),
            hashlib.sha256(content).hexdigest(),
        )
        for uploaded, content in zip(uploads, upload_bytes)
    )
    if st.session_state.get("tesco_input_signature") != signature:
        st.session_state.pop("tesco_generation_result", None)
        st.session_state["tesco_input_signature"] = signature

    if st.button("Generate reports", type="primary", use_container_width=True):
        try:
            with st.spinner("Adding new visits to the reports..."):
                st.session_state["tesco_generation_result"] = generate_reports(
                    export_bytes=upload_bytes[0],
                    calendar_bytes=upload_bytes[1],
                    main_report_bytes=upload_bytes[2],
                    whoosh_report_bytes=upload_bytes[3],
                    main_original_name=main_report_file.name,
                    whoosh_original_name=whoosh_report_file.name,
                )
        except (KeyError, ValueError, BadZipFile) as exc:
            st.session_state.pop("tesco_generation_result", None)
            st.error(str(exc).strip("'"))
        except Exception as exc:
            st.session_state.pop("tesco_generation_result", None)
            st.error(f"The reports could not be generated: {exc}")

    result = st.session_state.get("tesco_generation_result")
    if result is None:
        return

    st.success(f"Week {result.reporting_week} reports generated successfully.")
    summary_rows = []
    for audit_type in SUPPORTED_TYPES:
        summary_rows.append(
            {
                "Audit type": audit_type,
                "Added": result.stats[audit_type]["Added"],
                "Already present": result.stats[audit_type]["Already present"],
            }
        )
    st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)

    if result.ignored_rows:
        st.info(
            f"{result.ignored_rows} row(s) with other audit types were ignored."
        )

    st.download_button(
        "Download Test Purchase Report",
        data=result.main_report,
        file_name=result.main_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.download_button(
        "Download Whoosh Report",
        data=result.whoosh_report,
        file_name=result.whoosh_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.download_button(
        "Download both reports as ZIP",
        data=make_download_zip(result),
        file_name=f"Tesco Ireland Reports - Week {result.reporting_week}.zip",
        mime="application/zip",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
